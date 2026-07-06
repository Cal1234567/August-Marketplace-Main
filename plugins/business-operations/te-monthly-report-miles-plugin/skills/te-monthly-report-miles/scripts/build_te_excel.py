"""
build_te_excel.py — Builds the AG T&E Budget vs. Actual monthly Excel report.

Usage (CLI):
    python build_te_excel.py \
        --actuals '[{"department":"Advisory","acctnumber":"626100","actual_amount":1234.56}, ...]' \
        --budget '{"Advisory":{"626100":700.0},...}' \
        --month "May" \
        --year 2026 \
        --output /path/to/TE_BudgetVsActual_May2026.xlsx

Usage (Python import):
    from build_te_excel import build_report
    build_report(actuals_list, budget_dict, "May", 2026, "/path/to/output.xlsx")
"""

import argparse, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Account list (12 accounts matching template structure) ────────────────────
ACCOUNTS = [
    ("626100", "626100 - Accommodations"),
    ("626200", "626200 - Airline and Other Travel Tickets"),
    ("626300", "626300 - Car Mileage"),
    ("626400", "626400 - Car Rental"),
    ("626500", "626500 - Meals (Business - With Clients)"),
    ("626550", "626550 - Meals (No Clients)"),
    ("626560", "626560 - Tips/Gratuities"),
    ("626600", "626600 - Parking & Tolls"),
    ("626700", "626700 - Road & Rail (Uber, Taxi or Trains)"),
    ("626850", "626850 - Entertainment - August (No Clients)"),
    ("626900", "626900 - Entertainment - Client Events/Shows"),
    ("626910", "626910 - Entertainment - Client Golf"),
]
# NOTE: 626800 is excluded to match the 12-row template structure (rows 9-20 / 26-37).
# If 626800 data exists it will be captured in the Total column via no_dept lookup.

DEPARTMENTS = [
    "Advisory",
    "Business Development / Client Servicing",
    "Corporate Events",
    "Group Operations",
    "Investments",
    "Legal / Compliance",
]
SECTION1_DEPTS = DEPARTMENTS[:4]   # Advisory … Group Operations
SECTION2_DEPTS = DEPARTMENTS[4:]   # Investments, Legal / Compliance (+ Total)

GRAY = "FFD0D0D0"   # header fill
CURR = '"$"#,##0.00'


# ── Style helpers ─────────────────────────────────────────────────────────────

def _side(style="thin", color="000000"):
    return Side(style=style, color=color)

def _dot():
    """Dotted silver separator — used on last-data-row bottom and totals-row top."""
    return Side(style="dotted", color="C0C0C0")

def _border(l=None, r=None, t=None, b=None):
    return Border(
        left   = l or Side(style=None),
        right  = r or Side(style=None),
        top    = t or Side(style=None),
        bottom = b or Side(style=None),
    )

def _hfont(bold=True, sz=7):
    """Header / label font: Arial 7pt bold."""
    return Font(name="Arial", bold=bold, size=sz)

def _dfont(bold=False, sz=8):
    """Data / body font: Arial 8pt."""
    return Font(name="Arial", bold=bold, size=sz)

def _gray():
    return PatternFill("solid", fgColor=GRAY)


# ── Main builder ──────────────────────────────────────────────────────────────

def build_report(actuals_list, budget_input, month_name, year, output_path):
    """
    actuals_list : [{department, acctnumber, actual_amount}, ...]
    budget_input : nested  {dept: {acctnumber: amount}}  — preferred
                   OR flat {acctnumber: amount}           — consolidated fallback
    """
    # Normalise budget
    if budget_input and isinstance(list(budget_input.values())[0], dict):
        budget = budget_input
    else:
        budget = {"_consolidated": budget_input}

    def get_budget(dept, acct):
        if "_consolidated" in budget:
            return float(budget["_consolidated"].get(acct, 0) or 0)
        return float((budget.get(dept) or {}).get(acct, 0) or 0)

    # Build actuals lookup
    actual_lkp: dict = {}
    no_dept:    dict = {}
    for row in actuals_list:
        d = row["department"]
        a = row["acctnumber"]
        v = float(row.get("actual_amount", 0) or 0)
        if d == "No Department":
            no_dept[a] = no_dept.get(a, 0) + v
        else:
            actual_lkp[(a, d)] = actual_lkp.get((a, d), 0) + v

    wb = Workbook()
    ws = wb.active
    ws.title = f"{month_name} {year}"

    # ── Print / page setup ────────────────────────────────────────────────────
    ws.page_setup.orientation  = "landscape"
    ws.page_setup.paperSize    = 1          # Letter
    ws.page_setup.fitToPage    = True
    ws.page_setup.fitToWidth   = 1
    ws.page_setup.fitToHeight  = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # ── Row 1: blank merged (B1:Z1) ───────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=26)
    ws.row_dimensions[1].height = 15.6

    # ── Title rows 2-4 ────────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 15.6
    ws.row_dimensions[3].height = 17.4
    ws.row_dimensions[4].height = 17.4
    ws.row_dimensions[5].height = 17.4

    def _title(row, text, sz=12):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=14)
        c = ws.cell(row=row, column=2, value=text)
        c.font = Font(name="Arial", bold=True, size=sz)
        c.alignment = Alignment(horizontal="center", vertical="center")

    _title(2, "AG Reporting : August Group (Consolidated)", 12)
    _title(3, "Budget vs. Actual", 14)
    _title(4, f"{month_name} {year}", 14)

    # Row 5: bottom solid border only (divider beneath title block)
    for col in range(2, 15):   # B through N
        ws.cell(row=5, column=col).border = _border(b=_side())

    # ── Section writer ────────────────────────────────────────────────────────
    def write_section(start_row, depts, include_total=False):
        H1 = start_row              # dept name header row
        H2 = start_row + 1         # sub-column label row
        D0 = start_row + 2         # 626000 category header row
        DR = D0 + 1                # first account data row
        TR = DR + len(ACCOUNTS)    # totals row
        last_dr = TR - 1           # last account row (gets dotted bottom)
        n_depts = len(depts)
        n_groups = n_depts + (1 if include_total else 0)

        # Row heights for this entire section (10.2 per spec)
        for r in range(H1, TR + 1):
            ws.row_dimensions[r].height = 10.2

        # ── Column B: dept header ────────────────────────────────────────────
        if include_total:
            c = ws.cell(row=H1, column=2, value=None)
            c.font = _hfont()
            c.border = _border(l=_side(), r=_side(), t=_side())
        else:
            c = ws.cell(row=H1, column=2, value="Financial Row")
            c.font = _hfont()
            c.fill = _gray()
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = _border(l=_side(), r=_side(), t=_side())

        # ── Dept merged header cells ──────────────────────────────────────────
        for i, dept in enumerate(depts):
            cs = 3 + i * 3
            ws.merge_cells(start_row=H1, start_column=cs, end_row=H1, end_column=cs + 2)
            c = ws.cell(row=H1, column=cs, value=dept)
            c.font = _hfont(); c.fill = _gray()
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = _border(l=_side(), t=_side())
            # Middle cell: top only, no fill
            ws.cell(row=H1, column=cs + 1).border = _border(t=_side())
            # End cell: top + right (every group gets a right border)
            ws.cell(row=H1, column=cs + 2).border = _border(t=_side(), r=_side())

        # ── Total merged header (section 2 only) ─────────────────────────────
        if include_total:
            cs = 3 + n_depts * 3
            ws.merge_cells(start_row=H1, start_column=cs, end_row=H1, end_column=cs + 2)
            c = ws.cell(row=H1, column=cs, value="Total")
            c.font = _hfont(); c.fill = _gray()
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _border(l=_side(), r=_side(), t=_side())
            ws.cell(row=H1, column=cs + 1).border = _border(t=_side())
            ws.cell(row=H1, column=cs + 2).border = _border(t=_side(), r=_side())

        # ── Sub-column labels (H2) ────────────────────────────────────────────
        # Col B
        if include_total:
            c = ws.cell(row=H2, column=2, value=None)
            c.font = _hfont()
            c.border = _border(l=_side(), r=_side())
        else:
            c = ws.cell(row=H2, column=2, value="\xa0")
            c.font = _hfont(); c.fill = _gray()
            c.border = _border(l=_side(), r=_side())

        SUB = ["Amount", "Budget Amount", "Amount Over Budget"]
        for g in range(n_groups):
            for k, lbl in enumerate(SUB):
                col = 3 + g * 3 + k
                c = ws.cell(row=H2, column=col, value=lbl)
                c.font = _hfont(); c.fill = _gray()
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                # Amount col: left border; Budget: none; AOB: right border
                c.border = _border(
                    l=_side() if k == 0 else None,
                    r=_side() if k == 2 else None,
                )

        # ── 626000 category header ────────────────────────────────────────────
        c = ws.cell(row=D0, column=2, value="626000 - Travel and Entertainment")
        c.font = _dfont(bold=True)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = _border(l=_side(), r=_side())
        for g in range(n_groups):
            for k in range(3):
                col = 3 + g * 3 + k
                cell = ws.cell(row=D0, column=col)
                cell.font = _dfont(bold=True)
                cell.border = _border(
                    l=_side() if k == 0 else None,
                    r=_side() if k == 2 else None,
                )

        # ── Account data rows ─────────────────────────────────────────────────
        for ai, (acct, label) in enumerate(ACCOUNTS):
            r = DR + ai
            is_last = (r == last_dr)
            bot = _dot() if is_last else None  # dotted silver on last data row

            # Col B label
            c = ws.cell(row=r, column=2, value=label)
            c.font = _dfont()
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = _border(l=_side(), r=_side(), b=bot)

            # Per-dept columns
            for g, dept in enumerate(depts):
                ac = 3 + g * 3
                bc = ac + 1
                oc = ac + 2
                ar = f"{