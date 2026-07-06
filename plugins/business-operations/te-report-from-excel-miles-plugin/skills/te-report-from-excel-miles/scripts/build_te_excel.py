"""
build_te_excel.py — Builds the AG T&E Budget vs. Actual monthly Excel report.
"""
import argparse, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ACCOUNTS = [
    ("626100", "626100 - Accommodations"),
    ("626200", "626200 - Airline and Other Travel Tickets"),
    ("626300", "626300 - Car Mileage"),
    ("626400", "626400 - Car Rental"),
    ("626500", "626500 - Meals (Business - With Clients)"),
    ("626550", "626550 - Meals (No Clients)"),
    ("626560", "626560 - Tips/Gratuities"),
    ("626600", "626600 - Parking & Tolls"),
    ("626700", "626700 - Road & Rail (Uber, Taxi or Trains)"),
    ("626850", "626850 - Entertainment - August (No Clients)"),
    ("626900", "626900 - Entertainment - Client Events/Shows"),
    ("626910", "626910 - Entertainment - Client Golf"),
]

DEPARTMENTS = [
    "Advisory",
    "Business Development / Client Servicing",
    "Corporate Events",
    "Group Operations",
    "Investments",
    "Legal / Compliance",
]
SECTION1_DEPTS = DEPARTMENTS[:4]
SECTION2_DEPTS = DEPARTMENTS[4:]

GRAY = "FFD0D0D0"
CURR = '"$"#,##0.00'


def _side(style="thin", color="000000"):
    return Side(style=style, color=color)

def _dot():
    return Side(style="dotted", color="C0C0C0")

def _border(l=None, r=None, t=None, b=None):
    return Border(
        left   = l or Side(style=None),
        right  = r or Side(style=None),
        top    = t or Side(style=None),
        bottom = b or Side(style=None),
    )

def _hfont(bold=True, sz=7):
    return Font(name="Arial", bold=bold, size=sz)

def _dfont(bold=False, sz=8):
    return Font(name="Arial", bold=bold, size=sz)

def _gray_fill():
    return PatternFill("solid", fgColor=GRAY)


def build_report(actuals_list, budget_input, month_name, year, output_path):
    if budget_input and isinstance(list(budget_input.values())[0], dict):
        budget = budget_input
    else:
        budget = {"_consolidated": budget_input}

    def get_budget(dept, acct):
        if "_consolidated" in budget:
            return float(budget["_consolidated"].get(acct, 0) or 0)
        return float((budget.get(dept) or {}).get(acct, 0) or 0)

    actual_lkp = {}
    no_dept = {}
    for row in actuals_list:
        d = row["department"]
        a = row["acctnumber"]
        v = float(row.get("actual_amount", 0) or 0)
        if d == "No Department":
            no_dept[a] = no_dept.get(a, 0) + v
        else:
            actual_lkp[(a, d)] = actual_lkp.get((a, d), 0) + v

    wb = Workbook()
    ws = wb.active
    ws.title = f"{month_name} {year}"

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = 1
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.column_dimensions["B"].width = 38
    for col_idx in range(3, 15):
        pos = (col_idx - 3) % 3
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            19 if pos == 2 else (16 if pos == 1 else 14)
        )

    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=26)
    ws.row_dimensions[1].height = 15.6
    ws.row_dimensions[2].height = 15.6
    ws.row_dimensions[3].height = 17.4
    ws.row_dimensions[4].height = 17.4
    ws.row_dimensions[5].height = 17.4

    def _title(row, text, sz=12):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=14)
        c = ws.cell(row=row, column=2, value=text)
        c.font = Font(name="Arial", bold=True, size=sz)
        c.alignment = Alignment(horizontal="center", vertical="center")

    _title(2, "AG Reporting : August Group (Consolidated)", 12)
    _title(3, "Budget vs. Actual", 14)
    _title(4, f"{month_name} {year}", 14)

    for col in range(2, 15):
        ws.cell(row=5, column=col).border = _border(b=_side())

    def write_section(start_row, depts, include_total=False,
                      s1_act_refs=None, s1_bud_refs=None, s1_dr_start=None):
        H1 = start_row
        H2 = start_row + 1
        D0 = start_row + 2
        DR = D0 + 1
        TR = DR + len(ACCOUNTS)
        last_dr = TR - 1
        n_depts  = len(depts)
        n_groups = n_depts + (1 if include_total else 0)

        for r in range(H1, TR + 1):
            ws.row_dimensions[r].height = 10.2

        if include_total:
            c = ws.cell(row=H1, column=2, value=None)
            c.font = _hfont()
            c.border = _border(l=_side(), r=_side(), t=_side())
        else:
            c = ws.cell(row=H1, column=2, value="Financial Row")
            c.font = _hfont(); c.fill = _gray_fill()
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = _border(l=_side(), r=_side(), t=_side())

        for i, dept in enumerate(depts):
            cs = 3 + i * 3
            ws.merge_cells(start_row=H1, start_column=cs, end_row=H1, end_column=cs + 2)
            c = ws.cell(row=H1, column=cs, value=dept)
            c.font = _hfont(); c.fill = _gray_fill()
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = _border(l=_side(), t=_side())
            ws.cell(row=H1, column=cs + 1).border = _border(t=_side())
            ws.cell(row=H1, column=cs + 2).border = _border(t=_side(), r=_side())

        if include_total:
            cs = 3 + n_depts * 3
            ws.merge_cells(start_row=H1, start_column=cs, end_row=H1, end_column=cs + 2)
            c = ws.cell(row=H1, column=cs, value="Total")
            c.font = _hfont(); c.fill = _gray_fill()
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _border(l=_side(), r=_side(), t=_side())
            ws.cell(row=H1, column=cs + 1).border = _border(t=_side())
            ws.cell(row=H1, column=cs + 2).border = _border(t=_side(), r=_side())

        if include_total:
            c = ws.cell(row=H2, column=2, value=None)
            c.font = _hfont()
            c.border = _border(l=_side(), r=_side())
        else:
            c = ws.cell(row=H2, column=2, value="\xa0")
            c.font = _hfont(); c.fill = _gray_fill()
            c.border = _border(l=_side(), r=_side())

        SUB = ["Amount", "Budget Amount", "Amount Over Budget"]
        for g in range(n_groups):
            for k, lbl in enumerate(SUB):
                col = 3 + g * 3 + k
                c = ws.cell(row=H2, column=col, value=lbl)
                c.font = _hfont(); c.fill = _gray_fill()
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = _border(
                    l=_side() if k == 0 else None,
                    r=_side() if k == 2 else None,
                )

        c = ws.cell(row=D0, column=2, value="626000 - Travel and Entertainment")
        c.font = _dfont(bold=True)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = _border(l=_side(), r=_side())
        for g in range(n_groups):
            for k in range(3):
                col = 3 + g * 3 + k
                cell = ws.cell(row=D0, column=col)
                cell.font = _dfont(bold=True)
                cell.border = _border(
                    l=_side() if k == 0 else None,
                    r=_side() if k == 2 else None,
                )

        sect_act_letters = []
        sect_bud_letters = []

        for ai, (acct, label) in enumerate(ACCOUNTS):
            r = DR + ai
            is_last = (r == last_dr)
            bot = _dot() if is_last else None

            c = ws.cell(row=r, column=2, value=label)
            c.font = _dfont()
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = _border(l=_side(), r=_side(), b=bot)

            for g, dept in enumerate(depts):
                ac  = 3 + g * 3
                bc  = ac + 1
                oc  = ac + 2
                acl = get_column_letter(ac)
                bcl = get_column_letter(bc)
                ar  = acl + str(r)
                br  = bcl + str(r)

                act_val = actual_lkp.get((acct, dept), 0)
                bud_val = get_budget(dept, acct)

                cell = ws.cell(row=r, column=ac, value=act_val)
                cell.number_format = CURR; cell.font = _dfont()
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border = _border(l=_side(), b=bot)

                cell = ws.cell(row=r, column=bc, value=bud_val)
                cell.number_format = CURR; cell.font = _dfont()
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border = _border(b=bot)

                over_val = round(act_val, 2) - round(bud_val, 2)
                cell = ws.cell(row=r, column=oc, value=over_val)
                cell.number_format = CURR; cell.font = _dfont()
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border = _border(r=_side(), b=bot)

                if ai == 0:
                    sect_act_letters.append(acl)
                    sect_bud_letters.append(bcl)

            if include_total and s1_act_refs is not None:
                tot_ac = 3 + n_depts * 3
                tot_bc = tot_ac + 1
                tot_oc = tot_ac + 2
                tar_l  = get_column_letter(tot_ac)
                tbr_l  = get_column_letter(tot_bc)
                tar    = tar_l + str(r)
                tbr    = tbr_l + str(r)

                s1_row = s1_dr_start + ai
                s1_act_cells = [col + str(s1_row) for col in s1_act_refs]
                s1_bud_cells = [col + str(s1_row) for col in s1_bud_refs]
                s2_act_cells = [col + str(r) for col in sect_act_letters]
                s2_bud_cells = [col + str(r) for col in sect_bud_letters]
                nd = no_dept.get(acct, 0)
                all_act_cells = s1_act_cells + s2_act_cells
                all_bud_cells = s1_bud_cells + s2_bud_cells
                tot_act = "=SUM(" + ",".join(all_act_cells) + ")"
                if nd:
                    tot_act = "=SUM(" + ",".join(all_act_cells) + ")+" + str(nd)
                tot_bud = "=SUM(" + ",".join(all_bud_cells) + ")"

                _ta2 = sum(actual_lkp.get((acct, d), 0) for d in DEPARTMENTS)
                _ta2 += no_dept.get(acct, 0)
                _tb2 = sum(get_budget(d, acct) for d in DEPARTMENTS)
                cell = ws.cell(row=r, column=tot_ac, value=round(_ta2, 2))
                cell.number_format = CURR; cell.font = _dfont()
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border = _border(l=_side(), b=bot)

                cell = ws.cell(row=r, column=tot_bc, value=round(_tb2, 2))
                cell.number_format = CURR; cell.font = _dfont()
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border = _border(b=bot)

                # parse tot_act/tot_bud formula strings to get numeric totals
                _ta = sum(actual_lkp.get((acct, d), 0) for d in DEPARTMENTS[:4] + DEPARTMENTS[4:])
                _ta += no_dept.get(acct, 0)
                _tb = sum(get_budget(d, acct) for d in DEPARTMENTS)
                cell = ws.cell(row=r, column=tot_oc, value=round(_ta, 2) - round(_tb, 2))
                cell.number_format = CURR; cell.font = _dfont()
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border = _border(r=_side(), b=bot)

        # Totals row
        c = ws.cell(row=TR, column=2,
                    value="Total - 626000 - Travel and Entertainment")
        c.font = _dfont(bold=True)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = _border(l=_side(), r=_side(), t=_dot(), b=_side())

        for g in range(n_groups):
            ac  = 3 + g * 3
            bc  = ac + 1
            oc  = ac + 2
            acl = get_column_letter(ac)
            bcl = get_column_letter(bc)
            tar = acl + str(TR)
            tbr = bcl + str(TR)

            # Compute dept total from data rows
            if g < n_depts:
                dept = depts[g]
                tot_a = sum(actual_lkp.get((acct, dept), 0) for acct, _ in ACCOUNTS)
                tot_b = sum(get_budget(dept, acct) for acct, _ in ACCOUNTS)
            else:
                # Total column: all 6 depts + no_dept
                tot_a = sum(actual_lkp.get((acct, d), 0) for acct, _ in ACCOUNTS for d in DEPARTMENTS)
                tot_a += sum(no_dept.get(acct, 0) for acct, _ in ACCOUNTS)
                tot_b = sum(get_budget(d, acct) for acct, _ in ACCOUNTS for d in DEPARTMENTS)
            cell = ws.cell(row=TR, column=ac, value=round(tot_a, 2))
            cell.number_format = CURR; cell.font = _dfont(bold=True)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = _border(l=_side(), t=_dot(), b=_side())

            cell = ws.cell(row=TR, column=bc, value=round(tot_b, 2))
            cell.number_format = CURR; cell.font = _dfont(bold=True)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = _border(t=_dot(), b=_side())

            cell = ws.cell(row=TR, column=oc, value=round(tot_a, 2) - round(tot_b, 2))
            cell.number_format = CURR; cell.font = _dfont(bold=True)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = _border(r=_side(), t=_dot(), b=_side())

        return sect_act_letters, sect_bud_letters

    S1_START = 6
    s1_act, s1_bud = write_section(S1_START, SECTION1_DEPTS)

    S2_START = S1_START + 16 + 1
    ws.row_dimensions[S2_START - 1].height = 10.2

    s1_dr_start = S1_START + 3   # H1 + H2 + D0 offsets = data starts at +3
    write_section(S2_START, SECTION2_DEPTS, include_total=True,
                  s1_act_refs=s1_act, s1_bud_refs=s1_bud, s1_dr_start=s1_dr_start)

    wb.save(output_path)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--actuals", required=True)
    parser.add_argument("--budget",  required=True)
    parser.add_argument("--month",   required=True)
    parser.add_argument("--year",    required=True, type=int)
    parser.add_argument("--output",  required=True)
    args = parser.parse_args()
    build_report(
        json.loads(args.actuals),
        json.loads(args.budget),
        args.month,
        args.year,
        args.output,
    )
    print("Written:", args.output)
